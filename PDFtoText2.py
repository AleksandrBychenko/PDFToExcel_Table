import PyPDF2
import textract
import numpy as np
import openpyxl
import tabula


# Открываем PDF-файл
with open("54564-54566.pdf", 'rb') as file:
    # Создаем объект для чтения PDF-файла
    pdf_reader = PyPDF2.PdfReader(file)

    # Создаем массив для хранения текстовых данных
    text_data = []
    #@@@@@@@@@@@@@@@@@@@@@@@@@@
    # Открываем Excel-файл 
    #workbook = openpyxl.Workbook()
    import os
    path = "file.xlsx"

    if os.path.exists(path):
        # Открываем файл Excel
        workbook = openpyxl.load_workbook(path)
    else:
        # Создаем новый файл Excel
        workbook = openpyxl.Workbook()
        workbook.save(path)

    #workbook = openpyxl.load_workbook('file.xlsx')
    # Получаем активный лист
    worksheet = workbook.active

    if worksheet.title != 'Sheet1':
        worksheet = workbook.create_sheet('Sheet1')
   

    # Очищаем содержимое листа
    worksheet.delete_rows(1, worksheet.max_row)

    # Сохраняем изменения в файл Excel
    workbook.save(path)

    #workbook = openpyxl.load_workbook('file.xlsx')
    #@@@@@@@@@@@@@@2

    # Проходим по каждой странице PDF-файла
    #for page_num in range(len(pdf_reader.pages)):
    for page_num in range(1):
        # Получаем объект страницы
        page_num = 5
        page = pdf_reader.pages[page_num]

        # Извлекаем текстовые данные из страницы
        text = page.extract_text()

        # Делим текст на строки и сохраняем в массив
        lines = text.split('\n')

        

        #_______________>
        
        #считываем строчку 
        send = []

        send1 = ['CUSTOMER']
        send.append(send1)

        send2  = []
        import re
        ln1 = re.split('\s+', lines[1])
        send2.append(str(ln1[4])+ ' '+str(ln1[5]) + '\n' + str(lines[2]) + '\n' + str(lines[3]) + '\n' + str(lines[4]))
        send.append(send2)

        send3  = ['Cust Code', 'VAT NO', 'NO', 'Date', 'Page']
        send.append(send3)

        import re
        #send4 = re.split('\s+', lines[7])
        #print(lines[0], lines[1], lines[2], lines[3],lines[4], lines[5], lines[6],lines[7])
        send4 = []
        ln1 = re.split('\s+', lines[1])
        #print(lines[6].split(' ')[0], ln1, lines[0].split(' '))
        send4.append(lines[0].split(' ')[0])
        send4.append(lines[6].split(' ')[0])
        send4.append(ln1[1])
        send4.append(ln1[2])
        send4.append(ln1[3])

        
        send5  = ['ITEM Code', 'DESCRIPTION','SIZE', 'U.M', 'QTY', 'UNIT PRICE','VALUE','', 'C.I']
        send.append(send5)
        
        #send.append(send4)
        #print(send)

        #ДОБАВЛЯЕМ НУЖНУЮ ЧАСТЬ 

        import tabula
        import numpy as np
        import pandas as pd
        #x = tabula.read_pdf('54564-54566.pdf', stream = True, multiple_tables = False, pages= page_num, relative_area = True,area=(29.4, 0, 49.3 +29.4 ,78))
        x = tabula.read_pdf('54564-54566.pdf', stream = True, multiple_tables = False, pages= page_num, relative_area = True,area=(23, 0, 49.3 + 23 ,0 + 110))
        #x = tabula.read_pdf('54564-54566.pdf', stream = True, multiple_tables = False, pages= page_num, relative_area = True,area=(23, 0, 29.4,0 + 110))
        #x = tabula.read_pdf('54564-54566.pdf', stream = True, multiple_tables = False, pages= page_num, relative_area = True)
        y = tabula.read_pdf('54564-54566.pdf', stream = True, multiple_tables = False, pages= page_num, relative_area = True,area=(23, 0, 49.3 + 23 ,10))
        
        check = np.array(y)
        print(check)
        #print(check)
        send = np.array(x)
        df = pd.DataFrame(send[0])

 
        #НАЧИНАЕМ ДОБАВЛЯТЬ >>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>

        # Получаем активный лист
        worksheet = workbook.active

        # Создаем массив для новой строки
        new_row = ['Значение 1', 'Значение 2', 'Значение 3']

        # Добавляем новую строку в конец листа
        #for i in range(len(send)):
        #    worksheet.append(send[i])
        #   workbook.save('file.xlsx')
        

        worksheet.append(send1)

        #делаем жирную строку
        '''
        from openpyxl.styles import Font
        red_font = Font(bold=True)
       # Enumerate the cells in the second row
        for cell in worksheet["2:2"]:
            cell.font = red_font
        '''
        from openpyxl.styles import Font
        bold_font = Font(bold=True)

        for i in range(len(send1)) :
            worksheet.cell(row=1, column= i+1).font = bold_font

        worksheet.append(send2)

        worksheet.append(send3)
        #делать  шрифт жирным 
        from openpyxl.styles import Font
        bold_font = Font(bold=True)

        for i in range(len(send3)) :
            worksheet.cell(row=3, column= i+1).font = bold_font
        
 
        worksheet.append(send4)

        worksheet.append(send5)
        #делать  шрифт жирным 
        from openpyxl.styles import Font
        bold_font = Font(bold=True)

        for i in range(len(send5)) :
            worksheet.cell(row=5, column= i+1).font = bold_font

        # Сохраняем изменения в файл
        workbook.save('file.xlsx')
        
        workbook.close()
        #Добовляем таблицу 

        with pd.ExcelWriter("file.xlsx", mode="a", engine="openpyxl", if_sheet_exists='overlay') as writer:

            import openpyxl
            APIworkbook = openpyxl.load_workbook("file.xlsx")
            APISheet = APIworkbook['Sheet1']
            max_row_count = 0

            for row in APISheet.rows:
                for cell in row:
                    if cell.value:
                        max_row_count += 1
                        break

            df.to_excel(writer, startrow = 5, startcol= 0, header=False, index = False )
            
        '''
            for row in worksheet.iter_rows():

                for cell in row:

                    #first_word = cell.value.split(' ')[0]
                    print( cell.value)
                     # проверить значение ячейки на наличие слова "apple"
                    if cell.value == 'Country of origin: ITALY':
                         # изменить значение ячейки на "orange"
                        cell.value = 'orange'
                        # записать значение "banana" в следующую ячейку
                        next_cell = worksheet.cell(row=cell.row, column=cell.column+1)
                        next_cell.value = 'banana'
        '''
        # Открываем файл Excel
        workbook = openpyxl.load_workbook(path)
        worksheet = workbook.active
        #worksheet.append(send2)


            #workbook.save('file.xlsx')
        
            # ПРОХОДИМСЯ ПО ЭКСЕЛЮ  И  ИЗМЕНЯЕМ ЕГО  ЕСЛИ  ЧТО        
            # пройтись по всем ячейкам на листе
        for row in worksheet.iter_rows():

            for cell in row:

                #first_word = cell.value.split(' ')[0]
                #print( cell.value)
                 # проверить значение ячейки на наличие слова "apple"
                if  cell.value is not None and 'Country' in str(cell.value): 
                    
                        
                        # изменить значение ячейки на "orange"
                        my_list = cell.value.split('Country of origin:')
                        cell.value = 'Country of origin:'
                        worksheet.cell(row=cell.row, column=cell.column).font = bold_font
                        # записать значение "banana" в следующую ячейку
                        next_cell = worksheet.cell(row=cell.row, column=cell.column+1)

                        new_string = "".join(my_list)
                        my_list = new_string.split('H. S  :') 
                        next_cell.value = my_list[0]
                        
                
                for i in range(len(check)):
                    for j in range(len(check[i])):
                        for k in range(0, len(check[i][j]), 2):
                            if check[i][j][k] is not None and type(check[i][j][k]) == str and 'SOLE' in check[i][j][k]:
                                break
                            else:
                                if  cell.value is not None and type(check[i][j][k]) == str and check[i][j][k] in str(cell.value):
                                    my_list = cell.value.split(check[i][j][k])
                                    #print(my_list)
                                    next_cell = worksheet.cell(row=cell.row, column=cell.column+1)
                                    cell.value = check[i][j][k]
                                    next_cell.value = my_list[1]
                                    break
                



        #workbook.save('file.xlsx') 
        
        workbook.save('file.xlsx')
        # Закрываем файл Excel
        workbook.close()


        # Сохраняем изменения в файл
        #workbook.save('file.xlsx')
        



        # Добавляем текстовые данные в массив
        
        #text_data.append(send)

        
# Выводим массив с текстовыми данными
#print(text_data)