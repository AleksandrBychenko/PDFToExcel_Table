# work with exel 

from openpyxl import Workbook
import openpyxl


'''
wb = Workbook()
# делаем единственный лист активным 
ws = wb.active

ws1 = wb.create_sheet("Mysheet")

ws['A4'] = 5



wb.save("test.xlsx")
'''

"""
    Один из способов чтения значений
    :return:
    """
book = openpyxl.load_workbook(filename="ТЗ PREMIATA.xlsx")
sheet = book.active
    # sheet = book.worksheets[1]


    # for row in sheet.values:
    #     for cell in row:
    #         print(cell)

for row in sheet.iter_rows():
    for cell in row:
        print(cell.value)


