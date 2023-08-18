# import pdfplumber
# pdf = pdfplumber.open('50584-50590.pdf')
# page = pdf.pages[0]
# text = page.extract_text()
# print(text)
# pdf.close()

# # Function saving string to txt file
# file1=open(r"D:\VS projects\Python\PDFToText\\Test.txt","a")
# file1.writelines(text)

# Import the required Module
# import tabula
# # Read a PDF File
# df = tabula.read_pdf("50584-50590.pdf", pages='all')[0]
# # convert PDF into CSV
# tabula.convert_into("50584-50590.pdf", "test.xls", output_format="csv", pages='all')
# print(

import tabula
#from tabula.io import read_pdf
import pandas as pd
import numpy as np

column_names = ['ITEM CODE', 'DESCRIPTION', 'SIZE', 'U.M.','QTY', 'UNIT PRICE', '% DISC.', 'VALUE', 'C.I.']

#------------>
'''
PDF = tabula.read_pdf('Example.pdf', stream = True, multiple_tables = False)


#PDF = np.loadtxt("Example.txt", delimiter=" ")
#PDF2 = np.array(PDF)
#print(PDF2)

PDF = pd.DataFrame(np.reshape(PDF,(29,9)))


PDF.columns = column_names
PDF.to_excel('file_'+str(1)+'.xlsx')
print("Done")
'''

#<---------------


#________>
'''
from PyPDF2 import PdfReader
  
# creating a pdf reader object
reader = PdfReader('54564-54566.pdf')
  
# printing number of pages in pdf file
print(len(reader.pages))
  
# getting a specific page from the pdf file
page = reader.pages[0]
page2 = reader.pages[1]
  
# extracting text from page
text = page.extract_text()
text2 = page2.extract_text()

l = text.split('\n')
send  = []
send = np.append(send, l[1] + '\n' + l[2] + '\n' + l[3] + '\n' + l[4])
print(text2)

PDF = pd.DataFrame(send)
#PDF.columns = column_names
PDF.to_excel('file_'+str(1)+'.xlsx')
print("Done")
'''
#<____________



'''
myfile = 'Example.pdf'
tabula.convert_into(myfile, "marksheet.csv")
'''




#df[0].columns = column_names

'''
for i in range(len(df)):
    df.to_excel('file_'+str(1)+'.xlsx')
'''


# from tika import parser

# rawText = parser.from_file('50584-50590.pdf')

# rawList = rawText['content'].splitlines()
# a = 0


#------------>
'''
from openpyxl import Workbook


x = tabula.read_pdf('54564-54566.pdf', stream = True, multiple_tables = False, pages="all")


#PDF = np.loadtxt("Example.txt", delimiter=" ")
#PDF2 = np.array(PDF)
#print(PDF2)

for i in x:    #x values in list []
    print("printing all the table from the sheet", i)
    df = pd.DataFrame(i)
    a = np.array([1, 2, 3])
    #a  =  pd.DataFrame(a)
    #df.columns = column_name
    #df.to_excel('tables.xlsx', header=True, index = True)
    #a.to_excel('tables.xlsx', header=True, index = True)
    df.to_excel('tables.xlsx', header=True, index = True)
    wb = Workbook()

# grab the active worksheet
    ws = wb.active

# Data can be assigned directly to cells
    ws['A1'] = 42

# Rows can also be appended
    ws.append([1, 2, 3])

# Python types will automatically be converted

# Save the file
    #wb.save("tables.xlsx")
    

 #df.to_excel('tables.xlsx', header=True, index = True)
'''
#<---------------


#xfsvnjsd
#------------>

from openpyxl import Workbook
from PyPDF2 import PdfReader
from openpyxl import load_workbook

# creating a pdf reader object
reader = PdfReader('54564-54566.pdf')


'''
# printing number of pages in pdf file len(reader.pages)
for i in range(2):
    x = tabula.read_pdf('54564-54566.pdf', stream = True, multiple_tables = False, pages= 1,  pandas_options={'header': None}, relative_area = True,area=(0, 0, 590, 580))

    send = np.array(x)
    wb = load_workbook('tables.xlsx')
    ws = wb.active
    df = pd.DataFrame(send[0])
    #ws.append(df) 
    #wb.save('tables.xlsx')    
    wb.close()

    send = np.array(x)
    #send  = send.reshape(30*9,-1)
    #print (send)
    
    df = pd.DataFrame(send[0])
    #df.columns = column_names
    df.to_excel('tables.xlsx', header=False, index = True)

'''
#(y,x)
#x = tabula.read_pdf('54564-54566.pdf', stream = True, multiple_tables = False, pages= 1,  pandas_options={'header': None}, relative_area = True,area=(0, 0, 590, 580))
x = tabula.read_pdf('54564-54566.pdf', stream = True, multiple_tables = False, pages= 1, relative_area = True,area=(27.6, 0, 78, 10))

print(x)

send = np.array(x)
df = pd.DataFrame(send[0])
                  
#df.columns = column_names
df.to_excel('tables.xlsx', header=False, index = True)
print(send[0][1])



#PDF = np.loadtxt("Example.txt", delimiter=" ")
#PDF2 = np.array(PDF)
#print(PDF2)

'''
for i in x:    #x values in list []
    print("printing all the table from the sheet", i)
    df = pd.DataFrame(i)
    a = np.array([1, 2, 3])
    #a  =  pd.DataFrame(a)
    #df.columns = column_name
    #df.to_excel('tables.xlsx', header=True, index = True)
    #a.to_excel('tables.xlsx', header=True, index = True)
    df.to_excel('tables.xlsx', header=True, index = True)
    wb = Workbook()

# grab the active worksheet
    ws = wb.active

# Data can be assigned directly to cells
    ws['A1'] = 42

# Rows can also be appended
    ws.append([1, 2, 3])

# Python types will automatically be converted

# Save the file
    #wb.save("tables.xlsx")
    

 #df.to_excel('tables.xlsx', header=True, index = True)
'''
#<---------------